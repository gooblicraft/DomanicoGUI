from openpyxl import *
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def format_excel():
    path ="student_scores.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Where the magic happens
    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = max_length + 2

    workbook.save("student_scores.xlsx")